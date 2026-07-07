from __future__ import annotations

from io import BytesIO

import pandas as pd
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from config import EXPORT_COLUMNS


def _safe_sheet_name(name: str) -> str:
    return name[:31]


def build_excel(results: pd.DataFrame, kpis: dict[str, float], warnings: pd.DataFrame) -> bytes:
    output = BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl", datetime_format="yyyy-mm-dd") as writer:
        summary = pd.DataFrame(
            [
                {"KPI": "Total pedidos con quiebre", "Valor": kpis.get("total_pedidos", 0)},
                {"KPI": "Pedidos reasignables", "Valor": kpis.get("reasignables", 0)},
                {"KPI": "Pedidos no reasignables", "Valor": kpis.get("no_reasignables", 0)},
                {"KPI": "Pedidos parciales", "Valor": kpis.get("parciales", 0)},
                {"KPI": "Unidades en riesgo", "Valor": kpis.get("unidades_riesgo", 0)},
                {"KPI": "Unidades recuperables", "Valor": kpis.get("unidades_recuperables", 0)},
                {"KPI": "Recuperación posible %", "Valor": kpis.get("recuperacion_pct", 0) / 100},
                {"KPI": "Pedidos con riesgo D-1", "Valor": kpis.get("riesgo_d1", 0)},
                {"KPI": "Venta recuperable estimada", "Valor": kpis.get("venta_recuperable", 0)},
            ]
        )
        summary.to_excel(writer, index=False, sheet_name="Resumen KPIs")
        if not warnings.empty:
            warnings.to_excel(writer, index=False, sheet_name="Validaciones")

        exportable = results.copy()
        for column in EXPORT_COLUMNS:
            if column not in exportable.columns:
                exportable[column] = None
        exportable = exportable[EXPORT_COLUMNS + [c for c in exportable.columns if c not in EXPORT_COLUMNS]]

        sheets = {
            "Pedidos reasignables": exportable[exportable["Estado"].isin(["Reasignable", "Riesgo por stock D-1"])],
            "Pedidos no reasignables": exportable[exportable["Estado"].eq("No reasignable")],
            "Revisión necesaria": exportable[exportable["Estado"].isin(["Requiere revisión", "Reasignable parcial", "Riesgo por stock D-1"])],
            "Detalle por tienda": _detail_by_store(results),
            "Detalle por SKU": _detail_by_sku(results),
            "Detalle completo": exportable,
        }

        for sheet_name, dataframe in sheets.items():
            dataframe.to_excel(writer, index=False, sheet_name=_safe_sheet_name(sheet_name))

        _format_workbook(writer.book)

    output.seek(0)
    return output.read()


def _format_workbook(workbook) -> None:
    header_fill = PatternFill("solid", fgColor="111827")
    header_font = Font(color="FFFFFF", bold=True)

    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"
        max_row = worksheet.max_row
        max_column = worksheet.max_column
        if max_row and max_column:
            worksheet.auto_filter.ref = worksheet.dimensions
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
        for column_idx in range(1, max_column + 1):
            letter = get_column_letter(column_idx)
            worksheet.column_dimensions[letter].width = 18

    if "Resumen KPIs" in workbook.sheetnames:
        summary = workbook["Resumen KPIs"]
        summary.column_dimensions["A"].width = 34
        summary.column_dimensions["B"].width = 18
        if summary.max_row >= 7:
            summary["B7"].number_format = "0.0%"
        if summary.max_row >= 9:
            summary["B9"].number_format = '#,##0.00'


def _detail_by_store(results: pd.DataFrame) -> pd.DataFrame:
    if results.empty or "Tienda recomendada" not in results:
        return pd.DataFrame()
    data = results.dropna(subset=["Tienda recomendada"])
    if data.empty:
        return pd.DataFrame()
    return (
        data.groupby(["Tienda recomendada", "Estado"], dropna=False)
        .agg(Pedidos=("Número de pedido", "count"), Unidades=("Cantidad sugerida", "sum"), Venta=("Precio", "sum"))
        .reset_index()
        .sort_values(["Pedidos", "Unidades"], ascending=False)
    )


def _detail_by_sku(results: pd.DataFrame) -> pd.DataFrame:
    if results.empty:
        return pd.DataFrame()
    return (
        results.groupby(["SKU", "Estado"], dropna=False)
        .agg(Pedidos=("Número de pedido", "count"), Unidades=("Cantidad", "sum"), Recuperables=("Cantidad sugerida", "sum"))
        .reset_index()
        .sort_values(["Pedidos", "Unidades"], ascending=False)
    )
