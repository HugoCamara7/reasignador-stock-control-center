from __future__ import annotations

from io import BytesIO

import pandas as pd
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from reassignment_engine import NO_STOCK, REASSIGNABLE


def export_reassignment_excel(
    recommendations: pd.DataFrame,
    alternatives: pd.DataFrame,
    kpis: dict[str, float],
) -> bytes:
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl", datetime_format="yyyy-mm-dd") as writer:
        summary = pd.DataFrame(
            [
                {"KPI": "Total pedidos con quiebre", "Valor": kpis.get("total", 0)},
                {"KPI": "Pedidos reasignables", "Valor": kpis.get("reassignable", 0)},
                {"KPI": "Pedidos sin stock disponible", "Valor": kpis.get("no_stock", 0)},
                {"KPI": "Pedidos a revisar manualmente", "Valor": kpis.get("review", 0)},
                {"KPI": "Unidades en quiebre", "Valor": kpis.get("units_broken", 0)},
                {"KPI": "Unidades recuperables", "Valor": kpis.get("units_recoverable", 0)},
                {"KPI": "% recuperacion posible", "Valor": kpis.get("recovery_pct", 0) / 100},
            ]
        )
        summary.to_excel(writer, index=False, sheet_name="Resumen KPIs")

        recommendations[recommendations["Estado"].eq(REASSIGNABLE)].to_excel(
            writer, index=False, sheet_name="Pedidos reasignables"
        )
        recommendations[recommendations["Estado"].eq(NO_STOCK)].to_excel(
            writer, index=False, sheet_name="Sin stock disponible"
        )
        alternatives.to_excel(writer, index=False, sheet_name="Stock candidato por SKU")
        _top_model_color(recommendations).to_excel(writer, index=False, sheet_name="Top modelo color")
        recommendations.to_excel(writer, index=False, sheet_name="Detalle completo")

        _format_workbook(writer.book)

    output.seek(0)
    return output.read()


def _top_model_color(recommendations: pd.DataFrame) -> pd.DataFrame:
    if recommendations.empty:
        return pd.DataFrame()
    return (
        recommendations.groupby(["Modelo Color"], dropna=False)
        .agg(Pedidos=("Pedido", "count"), Unidades=("Cantidad", "sum"))
        .reset_index()
        .sort_values(["Pedidos", "Unidades"], ascending=False)
    )


def _format_workbook(workbook) -> None:
    header_fill = PatternFill("solid", fgColor="111827")
    header_font = Font(color="FFFFFF", bold=True)
    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"
        if worksheet.max_row and worksheet.max_column:
            worksheet.auto_filter.ref = worksheet.dimensions
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
        for column_idx in range(1, worksheet.max_column + 1):
            worksheet.column_dimensions[get_column_letter(column_idx)].width = 20
    if "Resumen KPIs" in workbook.sheetnames:
        workbook["Resumen KPIs"]["B7"].number_format = "0.0%"
